import schedule
from datetime import datetime
import openpyxl
import time

"""Этот скрипт должен запускаться отдельно, от всего проекта или в отдельном потоку"""
def greeting():
    global gauth
    cur_date = datetime.now().strftime("%d_%m_%Y")
    try:
        # если файл есть дописываем
        book = openpyxl.load_workbook("my_book.xlsx")
    except:
        # если нет создаем
        book = openpyxl.Workbook()

        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in book.sheetnames:
            sheet = book.get_sheet_by_name(sheet_name)
            book.remove_sheet(sheet)
            numbr = 1

    # создаем лист
    sheet = book.create_sheet(f'{cur_date}', 0)

    sheet['O1'].value = '.'  # чтобы лист создалось ставим точку в ячейкуу О
    print("создан новый лист")

    book.save("my_book.xlsx")
    book.close()

# запуск функции в каждый день в определенное время
schedule.every().days.at('00:01').do(greeting)
# schedule.every(1).minutes.do(greeting) # для теста

while True:
    time.sleep(20)
    print("Работаю")
    schedule.run_pending()