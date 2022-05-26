import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl import load_workbook
from datetime import datetime


# создаем еще один, так как в том нельзя делать взаимо импорты
edit2 = {'is': 0}
correct_number = {}


def loader(edit, chat_id, user_name, sp_phone, sp_tur, data_day, vz_sk, stoim_vz, posadoch, stoim_chi_1, besplat, naz_bes, nom_tel_tur, dop_inf, number=100):
    """сохраняем данные"""
    try:
        # если файл есть дописываем
        book = openpyxl.load_workbook("my_book.xlsx")
    except:
        # если нет создаем
        book = openpyxl.Workbook()
        cur_date = datetime.now().strftime("%d_%m_%Y")
        # Удаление листа, создаваемого по умолчанию, при создании документа
        for sheet_name in book.sheetnames:
            sheet = book.get_sheet_by_name(sheet_name)
            book.remove_sheet(sheet)

        # создаем новый лист
        sheet = book.create_sheet(f'{cur_date}', 0)


    # active берет первый лист
    sheet = book.active

    # создадим заголовки
    sheet['A1'] = 'НОМЕР ЗАЯВКИ'
    sheet['B1'] = 'ID'
    sheet['C1'] = 'АГЕНТ'
    sheet['D1'] = 'НОМЕР АГЕНТА'
    sheet['E1'] = 'ТУР'
    sheet['F1'] = 'ДАТА'
    sheet['G1'] = 'ВЗРОСЛЫЕ'
    sheet['H1'] = 'ЦЕНА ДЛЯ ВЗРОСЛЫХ'
    sheet['I1'] = 'ДЕТИ С ПОСАДОЧНЫМ МЕСТОМ'
    sheet['J1'] = 'СТОИМОСТЬ НА ОДНОГО РЕБЕНКА'
    sheet['K1'] = 'ДЕТИ БЕСПЛАТНО'
    sheet['l1'] = 'ОСТАНОВКА'
    sheet['M1'] = 'ТЕЛЕФОН ТУРИСТА'
    sheet['N1'] = 'ДОП. ИНФОРМАЦИЯ'

    # запишем данные
    # с пустого ряда
    row = 2  # начинаем поиск пустого ряда с ряда номер 2
    while True:
        if edit['is'] == 0:
            if sheet[f'A{row}'].value is None:
                sheet[f'A{row}'].value = number
                sheet[f'B{row}'].value = chat_id
                sheet[f'C{row}'].value = user_name
                sheet[f'D{row}'].value = sp_phone
                sheet[f'E{row}'].value = sp_tur
                sheet[f'F{row}'].value = data_day
                sheet[f'G{row}'].value = vz_sk
                sheet[f'H{row}'].value = stoim_vz
                sheet[f'I{row}'].value = posadoch
                sheet[f'J{row}'].value = stoim_chi_1
                sheet[f'K{row}'].value = besplat
                sheet[f'l{row}'].value = naz_bes
                sheet[f'M{row}'].value = nom_tel_tur
                sheet[f'N{row}'].value = dop_inf

                break

        if edit['is'] != 0:
            if sheet[f'A{row}'].value == int(edit['is']):
                # sheet[f'A{row}'].value = number номер заявки оставлям как есть
                sheet[f'B{row}'].value = chat_id
                sheet[f'C{row}'].value = user_name
                sheet[f'D{row}'].value = sp_phone
                sheet[f'E{row}'].value = sp_tur
                sheet[f'F{row}'].value = data_day
                sheet[f'G{row}'].value = vz_sk
                sheet[f'H{row}'].value = stoim_vz
                sheet[f'I{row}'].value = posadoch
                sheet[f'J{row}'].value = stoim_chi_1
                sheet[f'K{row}'].value = besplat
                sheet[f'l{row}'].value = naz_bes
                sheet[f'M{row}'].value = nom_tel_tur
                sheet[f'N{row}'].value = dop_inf


                # окрашиваем в голубой скоректированные
                sheet[f'A{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'B{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'C{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'D{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'E{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'F{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'G{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'H{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'I{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'J{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'K{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'l{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'M{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')
                sheet[f'N{row}'].fill = PatternFill(fill_type='solid', start_color='5A9FFF')

                break


        row += 1

    book.save("my_book.xlsx")
    book.close()

    return 'ЗАЯВКА ПЕРЕДАНА ДИСПЕТЧЕРУ'


def write_zayav(number, id):
    global perzap, correct_number
    correct_number['cor'] = number
    """Поиск заявки"""
    try:
        book = openpyxl.open("my_book.xlsx", read_only=True)
    except:
        return 'Заявок еще нет!'

    # берем последний лист
    sheet = book.active

    # ищем заявку клиента
    # должно соответствовать и номер звявки и его id
    # ходим по всем заявкам ищя номер заявки
    for row in range(2, sheet.max_row + 1):
        if int(sheet[f'A{row}'].value) == int(number):
            number = sheet[f'A{row}'].value
            # когда нашли номер заявки ищем id пльзователя и смотрим его ли
            if sheet[f'B{row}'].value == id:
                chat_id = sheet[f'B{row}'].value
                # когда находим собираем все данные с этого ряда
                user_name = sheet[f'C{row}'].value
                sp_phone = sheet[f'D{row}'].value
                sp_tur = sheet[f'E{row}'].value
                data_day = sheet[f'F{row}'].value
                vz_sk = sheet[f'G{row}'].value
                stoim_vz = sheet[f'H{row}'].value
                posadoch = sheet[f'I{row}'].value
                stoim_chi_1 = sheet[f'J{row}'].value
                besplat = sheet[f'K{row}'].value
                naz_bes = sheet[f'L{row}'].value
                nom_tel_tur = sheet[f'M{row}'].value
                dop_inf = sheet[f'N{row}'].value

                return user_name, sp_phone, sp_tur, data_day, vz_sk, stoim_vz, posadoch, stoim_chi_1, besplat, naz_bes, nom_tel_tur, dop_inf

            else:
                return "Это не ваша заявка!"


# отмена заявки и окрашивание в красный цвет
def delete_z(chat_id, number_otm):

    book = openpyxl.load_workbook("my_book.xlsx")

    sheet = book.active

    # ищем заявку клиента
    for row in range(2, sheet.max_row + 1):
        if sheet[f'A{row}'].value == int(number_otm):
            # когда нашли номер заявки ищем id пльзователя и смотрим его ли
            if sheet[f'B{row}'].value == chat_id:
                # окрашиваем в красный отмена
                sheet[f'A{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'B{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'C{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'D{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'E{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'F{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'G{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'H{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'I{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'J{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'K{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'l{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'M{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')
                sheet[f'N{row}'].fill = PatternFill(fill_type='solid', start_color='FF1E2D')

        book.save("my_book.xlsx")
        book.close()

    return "Ваша заявка отменена!"


